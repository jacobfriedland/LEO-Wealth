"""
Inserts XLOOKUP formulas into the Fiduciary sheet column F.
Each formula looks up the ISIN in column C against the appropriate VO sheet range
and returns the matching weight, or "NA" if not found.

Patches the saved file's XML to remove openpyxl's unwanted _xlfn.SINGLE() wrapper
so formulas work correctly in modern Excel (no @ prefix).

Usage: python insert_xlookup_formulas.py
"""

import openpyxl
import zipfile
import os
import re

FILE   = 'c:/Users/JacobFriedland/OneDrive - Leo Wealth/Claude/Master Models Drifted Weights Template.xlsm'
OUTPUT = 'c:/Users/JacobFriedland/OneDrive - Leo Wealth/Claude/Master Models Drifted Weights Template - Updated.xlsm'
TEMP   = OUTPUT + '.tmp.xlsm'

def xlookup(lookup_cell, sheet, isin_range, weight_range):
    return f'=XLOOKUP({lookup_cell},\'{sheet}\'!{isin_range},\'{sheet}\'!{weight_range},"NA")'

# Mapping: fiduciary_portfolio_name -> formula function (takes row number)
# To add EUR/AUD VO later, add entries here following the same pattern.
PORTFOLIO_FORMULAS = {
    # USD VO: Global ETF (rows 3-20, ISIN=H, Eq=Q, Bal=S, Mod=U, Cons=W, FI=Y)
    'Global Equity ETF':                lambda r: xlookup(f'C{r}', 'USD VO', 'H3:H20',    'Q3:Q20'),
    'Global Balanced ETF':              lambda r: xlookup(f'C{r}', 'USD VO', 'H3:H20',    'S3:S20'),
    'Global Moderate ETF':              lambda r: xlookup(f'C{r}', 'USD VO', 'H3:H20',    'U3:U20'),
    'Global Conservative ETF':          lambda r: xlookup(f'C{r}', 'USD VO', 'H3:H20',    'W3:W20'),
    'Global Fixed Income ETF':          lambda r: xlookup(f'C{r}', 'USD VO', 'H3:H20',    'Y3:Y20'),

    # USD VO: Global ETF (US Tax) (rows 43-62)
    'Global Equity ETF (US Tax)':       lambda r: xlookup(f'C{r}', 'USD VO', 'H43:H62',   'Q43:Q62'),
    'Global Balanced ETF (US Tax)':     lambda r: xlookup(f'C{r}', 'USD VO', 'H43:H62',   'S43:S62'),
    'Global Moderate ETF (US Tax)':     lambda r: xlookup(f'C{r}', 'USD VO', 'H43:H62',   'U43:U62'),
    'Global Fixed Income ETF (US Tax)': lambda r: xlookup(f'C{r}', 'USD VO', 'H43:H62',   'Y43:Y62'),

    # USD VO: US Focused ETF (rows 84-102, Bal=S)
    'US Focused Balanced ETF':          lambda r: xlookup(f'C{r}', 'USD VO', 'H84:H102',  'S84:S102'),

    # USD VO: USD Cash Management (rows 326-332)
    'USD Cash Management':              lambda r: xlookup(f'C{r}', 'USD VO', 'H326:H332', 'Q326:Q332'),

    # USD VO: USD Cash Management (US Tax) (rows 344-349)
    'USD Cash Management (US Tax)':     lambda r: xlookup(f'C{r}', 'USD VO', 'H344:H349', 'Q344:Q349'),

    # USD VO: MUNI CEF (rows 308-313)
    'Muni CEF':                         lambda r: xlookup(f'C{r}', 'USD VO', 'H308:H313', 'Q308:Q313'),

    # USD VO: Commodities variants
    'Commodities':                      lambda r: xlookup(f'C{r}', 'USD VO', 'H124:H133', 'Q124:Q133'),
    'Commodities (US Tax)':             lambda r: xlookup(f'C{r}', 'USD VO', 'H145:H154', 'Q145:Q154'),
    'Commodities Ex-Crypto':            lambda r: xlookup(f'C{r}', 'USD VO', 'H166:H173', 'Q166:Q173'),
    'Commodities Ex-Crypto (US Tax)':   lambda r: xlookup(f'C{r}', 'USD VO', 'H185:H192', 'Q185:Q192'),

    # USD VO: Asian / Crypto ETFs
    'Asian ETF':                        lambda r: xlookup(f'C{r}', 'USD VO', 'H267:H275', 'Q267:Q275'),
    'Asian ETF (US Tax)':               lambda r: xlookup(f'C{r}', 'USD VO', 'H288:H295', 'Q288:Q295'),
    'Crypto ETF':                       lambda r: xlookup(f'C{r}', 'USD VO', 'H249:H254', 'Q249:Q254'),

    # GBP VO (rows 4-19, ISIN=H, Eq=Q, Bal=S, Mod=U, Cons=W)
    'GBP Global Equity ETF':            lambda r: xlookup(f'C{r}', 'GBP VO', 'H4:H19',   'Q4:Q19'),
    'GBP Global Balanced ETF':          lambda r: xlookup(f'C{r}', 'GBP VO', 'H4:H19',   'S4:S19'),
    'GBP Global Moderate ETF':          lambda r: xlookup(f'C{r}', 'GBP VO', 'H4:H19',   'U4:U19'),
    'GBP Global Conservative ETF':      lambda r: xlookup(f'C{r}', 'GBP VO', 'H4:H19',   'W4:W19'),

    # Single Stock VO (data rows 6+, each portfolio has its own ISIN and Weight columns)
    'Global Brands':        lambda r: xlookup(f'C{r}', 'Single Stock VO', 'R6:R41',   'Z6:Z41'),
    'Global Healthcare':    lambda r: xlookup(f'C{r}', 'Single Stock VO', 'AS6:AS25', 'BA6:BA25'),
    'Japan REIT':           lambda r: xlookup(f'C{r}', 'Single Stock VO', 'BT6:BT20', 'CB6:CB20'),
    'Global REIT':          lambda r: xlookup(f'C{r}', 'Single Stock VO', 'CU6:CU24', 'DC6:DC24'),
    'UK Equity Income':     lambda r: xlookup(f'C{r}', 'Single Stock VO', 'DV6:DV25', 'ED6:ED25'),
    'US Brands':            lambda r: xlookup(f'C{r}', 'Single Stock VO', 'EW6:EW34', 'FE6:FE34'),
    'Global Technology':    lambda r: xlookup(f'C{r}', 'Single Stock VO', 'FX6:FX35', 'GF6:GF35'),
    'Asian Leaders':        lambda r: xlookup(f'C{r}', 'Single Stock VO', 'GY6:GY39', 'HG6:HG39'),
    'HK Equity Income':     lambda r: xlookup(f'C{r}', 'Single Stock VO', 'HZ6:HZ27', 'IH6:IH27'),
    'AI & Robotics':        lambda r: xlookup(f'C{r}', 'Single Stock VO', 'JA6:JA32', 'JI6:JI32'),
    'Sustainable Future':   lambda r: xlookup(f'C{r}', 'Single Stock VO', 'KB6:KB39', 'KJ6:KJ39'),
    'US REIT':              lambda r: xlookup(f'C{r}', 'Single Stock VO', 'LC6:LC20', 'LK6:LK20'),
    'Global Equity Income': lambda r: xlookup(f'C{r}', 'Single Stock VO', 'MD6:MD39', 'ML6:ML39'),
    'US Equity Income':     lambda r: xlookup(f'C{r}', 'Single Stock VO', 'NE6:NE31', 'NM6:NM31'),
    'Europe Equity Income': lambda r: xlookup(f'C{r}', 'Single Stock VO', 'OF6:OF32', 'ON6:ON32'),
    'Japan Equity Income':  lambda r: xlookup(f'C{r}', 'Single Stock VO', 'PG6:PG28', 'PO6:PO28'),
    'Cybersecurity':        lambda r: xlookup(f'C{r}', 'Single Stock VO', 'QH6:QH26', 'QP6:QP26'),

    # TODO: Add EUR VO and AUD VO mappings here when ready
    # 'EUR Global Equity ETF':   lambda r: xlookup(f'C{r}', 'EUR VO', 'H?:H?', 'Q?:Q?'),
    # 'AUD Global Equity ETF':   lambda r: xlookup(f'C{r}', 'AUD VO', '??:??', '??:??'),
}

# ---- Load source and write formulas ----
wb = openpyxl.load_workbook(FILE, keep_vba=True)
ws_fid = wb['Fiduciary']
ws_fid.cell(row=1, column=6, value='VO Weight')

# Find portfolio header rows (col A has name, col C is empty)
fid_blocks = []
for i, row in enumerate(ws_fid.iter_rows(max_row=ws_fid.max_row), 1):
    if i == 1:
        continue
    vals = [c.value for c in row]
    if vals[0] is not None and (vals[2] is None or str(vals[2]).strip() == ''):
        fid_blocks.append((i, str(vals[0]).strip()))

block_ranges = []
for idx, (row_num, name) in enumerate(fid_blocks):
    next_row = fid_blocks[idx+1][0] if idx+1 < len(fid_blocks) else ws_fid.max_row + 1
    block_ranges.append((row_num, next_row - 1, name))

formula_count = 0
na_count = 0

for start_row, end_row, portfolio_name in block_ranges:
    formula_fn = PORTFOLIO_FORMULAS.get(portfolio_name)
    for row_num in range(start_row + 1, end_row + 1):
        isin_val = ws_fid.cell(row=row_num, column=3).value
        if not isin_val or str(isin_val).strip() == '':
            continue
        if formula_fn:
            ws_fid.cell(row=row_num, column=6, value=formula_fn(row_num))
            formula_count += 1
        else:
            ws_fid.cell(row=row_num, column=6, value='NA')
            na_count += 1

wb.save(TEMP)
print(f'Formulas written: {formula_count}, NA: {na_count}')

# ---- Patch XML: remove openpyxl's _xlfn.SINGLE() wrapper ----
# openpyxl wraps XLOOKUP in _xlfn.SINGLE(...) which renders as @XLOOKUP in Excel.
# We strip that wrapper so Excel sees a plain XLOOKUP formula.
with zipfile.ZipFile(TEMP, 'r') as zin, zipfile.ZipFile(OUTPUT, 'w', zipfile.ZIP_DEFLATED) as zout:
    for item in zin.infolist():
        data = zin.read(item.filename)
        if item.filename.startswith('xl/worksheets/') and item.filename.endswith('.xml'):
            text = data.decode('utf-8')
            text = re.sub(
                r'_xlfn\.SINGLE\((_xlfn\.XLOOKUP\([^)]*(?:\([^)]*\)[^)]*)*\))\)',
                r'\1', text
            )
            text = text.replace('_xlfn.XLOOKUP', 'XLOOKUP')
            data = text.encode('utf-8')
        zout.writestr(item, data)

os.remove(TEMP)
print(f'Saved: {OUTPUT}')
